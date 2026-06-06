import openpyxl

wb = openpyxl.load_workbook("1-4-2026 TO 30-4-2026.xlsx")
ws = wb.active

for r in range(40, 101):
    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v is not None for v in row_vals):
        print(f"Row {r}: {row_vals}")
