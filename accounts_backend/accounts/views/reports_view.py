# reports_view.py
import calendar
import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from django.http import HttpResponse
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.views import APIView

from accounts.common.db        import query_all, query_one
from accounts.common.auth      import JWTAuthentication
from accounts.common.responses import common_response, StatusCode
from accounts.common.messages  import get_message
from accounts.serializers.reports_serializer import (
    GSTSummarySerializer,
    GSTR1Serializer,
    HSNSummarySerializer,
    MonthlySalesSerializer,
)


# ── Shared queries ────────────────────────────────────────────────────────────

def _fetch_gstr1(company_id: int, month: int, year: int) -> list:
    if month == 0:
        return query_all(
            """
            SELECT
                i.invoice_number,
                i.invoice_date,
                c.customer_name,
                c.gstin  AS customer_gstin,
                c.state  AS customer_state,
                i.grand_total,
                CASE WHEN LOWER(COALESCE(c.state, '')) = LOWER(COALESCE(comp.state, '')) THEN 'Local' ELSE 'Central' END AS local_central,
                'Inventory' AS invoice_type,
                ii.hsn_code,
                SUM(ii.quantity) AS quantity,
                SUM(ii.total_amount) AS amount,
                SUM(ii.taxable_amount) AS taxable_value,
                CASE WHEN i.igst_amount = 0 THEN ii.gst_percentage / 2.0 ELSE 0.0 END AS sgst_pct,
                SUM(ii.sgst_amount) AS sgst_amount,
                CASE WHEN i.igst_amount = 0 THEN ii.gst_percentage / 2.0 ELSE 0.0 END AS cgst_pct,
                SUM(ii.cgst_amount) AS cgst_amount,
                CASE WHEN i.igst_amount > 0 THEN ii.gst_percentage ELSE 0.0 END AS igst_pct,
                SUM(ii.igst_amount) AS igst_amount,
                0.0 AS cess,
                SUM(ii.cgst_amount + ii.sgst_amount + ii.igst_amount) AS total_gst
            FROM   invoice_items ii
            JOIN   invoices i ON i.invoice_id = ii.invoice_id
            JOIN   customers c ON c.customer_id = i.customer_id
            JOIN   company comp ON comp.company_id = i.company_id
            WHERE  i.company_id   = %s
              AND  i.invoice_type = 'TAX'
              AND  i.status       = 'A'
              AND  EXTRACT(YEAR  FROM i.invoice_date) = %s
            GROUP BY
                i.invoice_number,
                i.invoice_date,
                c.customer_name,
                c.gstin,
                c.state,
                i.grand_total,
                comp.state,
                ii.hsn_code,
                ii.gst_percentage,
                i.igst_amount
            ORDER BY i.invoice_date, i.invoice_number, ii.hsn_code
            """,
            (company_id, year)
        )
    else:
        return query_all(
            """
            SELECT
                i.invoice_number,
                i.invoice_date,
                c.customer_name,
                c.gstin  AS customer_gstin,
                c.state  AS customer_state,
                i.grand_total,
                CASE WHEN LOWER(COALESCE(c.state, '')) = LOWER(COALESCE(comp.state, '')) THEN 'Local' ELSE 'Central' END AS local_central,
                'Inventory' AS invoice_type,
                ii.hsn_code,
                SUM(ii.quantity) AS quantity,
                SUM(ii.total_amount) AS amount,
                SUM(ii.taxable_amount) AS taxable_value,
                CASE WHEN i.igst_amount = 0 THEN ii.gst_percentage / 2.0 ELSE 0.0 END AS sgst_pct,
                SUM(ii.sgst_amount) AS sgst_amount,
                CASE WHEN i.igst_amount = 0 THEN ii.gst_percentage / 2.0 ELSE 0.0 END AS cgst_pct,
                SUM(ii.cgst_amount) AS cgst_amount,
                CASE WHEN i.igst_amount > 0 THEN ii.gst_percentage ELSE 0.0 END AS igst_pct,
                SUM(ii.igst_amount) AS igst_amount,
                0.0 AS cess,
                SUM(ii.cgst_amount + ii.sgst_amount + ii.igst_amount) AS total_gst
            FROM   invoice_items ii
            JOIN   invoices i ON i.invoice_id = ii.invoice_id
            JOIN   customers c ON c.customer_id = i.customer_id
            JOIN   company comp ON comp.company_id = i.company_id
            WHERE  i.company_id   = %s
              AND  i.invoice_type = 'TAX'
              AND  i.status       = 'A'
              AND  EXTRACT(MONTH FROM i.invoice_date) = %s
              AND  EXTRACT(YEAR  FROM i.invoice_date) = %s
            GROUP BY
                i.invoice_number,
                i.invoice_date,
                c.customer_name,
                c.gstin,
                c.state,
                i.grand_total,
                comp.state,
                ii.hsn_code,
                ii.gst_percentage,
                i.igst_amount
            ORDER BY i.invoice_date, i.invoice_number, ii.hsn_code
            """,
            (company_id, month, year)
        )


def _fetch_hsn_summary(company_id: int, month: int, year: int) -> list:
    if month == 0:
        return query_all(
            """
            SELECT
                COALESCE(ii.hsn_code, p.hsn_code, 'N/A') AS hsn_code,
                SUM(ii.quantity)       AS total_qty,
                SUM(ii.taxable_amount) AS taxable_value,
                SUM(ii.cgst_amount)    AS cgst,
                SUM(ii.sgst_amount)    AS sgst,
                SUM(ii.igst_amount)    AS igst,
                SUM(ii.total_amount)   AS total
            FROM   invoice_items ii
            JOIN   invoices i ON i.invoice_id = ii.invoice_id
            LEFT   JOIN products p ON p.product_id = ii.product_id
            WHERE  i.company_id   = %s
              AND  i.invoice_type = 'TAX'
              AND  i.status       = 'A'
              AND  EXTRACT(YEAR  FROM i.invoice_date) = %s
            GROUP  BY COALESCE(ii.hsn_code, p.hsn_code, 'N/A')
            ORDER  BY taxable_value DESC
            """,
            (company_id, year)
        )
    else:
        return query_all(
            """
            SELECT
                COALESCE(ii.hsn_code, p.hsn_code, 'N/A') AS hsn_code,
                SUM(ii.quantity)       AS total_qty,
                SUM(ii.taxable_amount) AS taxable_value,
                SUM(ii.cgst_amount)    AS cgst,
                SUM(ii.sgst_amount)    AS sgst,
                SUM(ii.igst_amount)    AS igst,
                SUM(ii.total_amount)   AS total
            FROM   invoice_items ii
            JOIN   invoices i ON i.invoice_id = ii.invoice_id
            LEFT   JOIN products p ON p.product_id = ii.product_id
            WHERE  i.company_id   = %s
              AND  i.invoice_type = 'TAX'
              AND  i.status       = 'A'
              AND  EXTRACT(MONTH FROM i.invoice_date) = %s
              AND  EXTRACT(YEAR  FROM i.invoice_date) = %s
            GROUP  BY COALESCE(ii.hsn_code, p.hsn_code, 'N/A')
            ORDER  BY taxable_value DESC
            """,
            (company_id, month, year)
        )


def _validate_month_year(request: Request):
    """Returns (month, year, error_response). error_response is None if valid."""
    month = request.query_params.get('month')
    year  = request.query_params.get('year')
    if not month or not year:
        return None, None, common_response(
            StatusCode.BAD_REQUEST.value,
            "month and year are required."
        )
    try:
        return int(month), int(year), None
    except ValueError:
        return None, None, common_response(
            StatusCode.BAD_REQUEST.value,
            "month and year must be integers."
        )


# ── Views ─────────────────────────────────────────────────────────────────────

class GSTSummaryView(generics.GenericAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    serializer_class       = GSTSummarySerializer

    def get(self, request: Request):
        month, year, err = _validate_month_year(request)
        if err:
            return err

        company_id = request.user.company_id
        if month == 0:
            row = query_one(
                """
                SELECT
                    COUNT(*)::int                    AS invoice_count,
                    COALESCE(SUM(subtotal),      0)  AS taxable_amount,
                    COALESCE(SUM(cgst_amount),   0)  AS total_cgst,
                    COALESCE(SUM(sgst_amount),   0)  AS total_sgst,
                    COALESCE(SUM(igst_amount),   0)  AS total_igst,
                    COALESCE(SUM(grand_total),   0)  AS grand_total
                FROM   invoices
                WHERE  company_id   = %s
                  AND  invoice_type = 'TAX'
                  AND  status       = 'A'
                  AND  EXTRACT(YEAR  FROM invoice_date) = %s
                """,
                (company_id, year)
            )
        else:
            row = query_one(
                """
                SELECT
                    COUNT(*)::int                    AS invoice_count,
                    COALESCE(SUM(subtotal),      0)  AS taxable_amount,
                    COALESCE(SUM(cgst_amount),   0)  AS total_cgst,
                    COALESCE(SUM(sgst_amount),   0)  AS total_sgst,
                    COALESCE(SUM(igst_amount),   0)  AS total_igst,
                    COALESCE(SUM(grand_total),   0)  AS grand_total
                FROM   invoices
                WHERE  company_id   = %s
                  AND  invoice_type = 'TAX'
                  AND  status       = 'A'
                  AND  EXTRACT(MONTH FROM invoice_date) = %s
                  AND  EXTRACT(YEAR  FROM invoice_date) = %s
                """,
                (company_id, month, year)
            )

        return common_response(
            StatusCode.OK.value,
            get_message("GST summary fetched successfully"),
            self.get_serializer(row).data
        )


class GSTR1View(generics.GenericAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    serializer_class       = GSTR1Serializer

    def get(self, request: Request):
        month, year, err = _validate_month_year(request)
        if err:
            return err

        rows = _fetch_gstr1(request.user.company_id, month, year)
        serializer = self.get_serializer(rows, many=True)
        return common_response(
            StatusCode.OK.value,
            get_message("GSTR-1 report fetched successfully"),
            {
                'count': len(serializer.data),
                'gstr1': serializer.data,
            }
        )


class HSNSummaryView(generics.GenericAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    serializer_class       = HSNSummarySerializer

    def get(self, request: Request):
        month, year, err = _validate_month_year(request)
        if err:
            return err

        rows = _fetch_hsn_summary(request.user.company_id, month, year)
        serializer = self.get_serializer(rows, many=True)
        return common_response(
            StatusCode.OK.value,
            get_message("HSN summary fetched successfully"),
            {'hsn_summary': serializer.data}
        )


class MonthlySalesView(generics.GenericAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]
    serializer_class       = MonthlySalesSerializer

    def get(self, request: Request):
        company_id = request.user.company_id
        rows = query_all(
            """
            SELECT
                TO_CHAR(invoice_date, 'Mon YYYY')      AS month_label,
                EXTRACT(YEAR  FROM invoice_date)::int  AS yr,
                EXTRACT(MONTH FROM invoice_date)::int  AS mo,
                COALESCE(SUM(grand_total), 0)          AS total_sales,
                COALESCE(SUM(cgst_amount + sgst_amount + igst_amount), 0) AS total_tax,
                COUNT(*)::int                          AS invoice_count
            FROM   invoices
            WHERE  company_id = %s
              AND  status     = 'A'
              AND  invoice_date >= CURRENT_DATE - INTERVAL '12 months'
            GROUP  BY month_label, yr, mo
            ORDER  BY yr, mo
            """,
            (company_id,)
        )
        serializer = self.get_serializer(rows, many=True)
        return common_response(
            StatusCode.OK.value,
            get_message("Monthly sales trend fetched successfully"),
            {'monthly_sales': serializer.data}
        )


class GSTR1ExcelDownloadView(APIView):
    """
    GET /reports/gstr1/download/?month=6&year=2026
    Returns a downloadable .xlsx with two sheets:
      Sheet 1 — Invoice-wise GSTR-1 detail
      Sheet 2 — HSN-wise summary
    """
    authentication_classes = [JWTAuthentication]
    permission_classes     = [IsAuthenticated]

    # ── Style constants ───────────────────────────────────────────────────────
    _HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    _SECTION_FILL  = PatternFill("solid", fgColor="2E75B6")
    _TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")
    _HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
    _SECTION_FONT  = Font(bold=True, color="FFFFFF", size=10)
    _TOTAL_FONT    = Font(bold=True, size=10)
    _LABEL_FONT    = Font(bold=True, size=10)
    _CENTER        = Alignment(horizontal="center", vertical="center")
    _RIGHT         = Alignment(horizontal="right",  vertical="center")

    def get(self, request: Request):
        month, year, err = _validate_month_year(request)
        if err:
            return err

        company_id = request.user.company_id

        company = query_one(
            "SELECT company_name, gstin FROM company WHERE company_id = %s",
            (company_id,)
        ) or {}

        gstr1_rows = _fetch_gstr1(company_id, month, year)
        hsn_rows   = _fetch_hsn_summary(company_id, month, year)

        wb = openpyxl.Workbook()

        self._build_gstr1_sheet(wb.active, gstr1_rows, company, month, year)
        hsn_ws = wb.create_sheet("HSN Summary")
        self._build_hsn_sheet(hsn_ws, hsn_rows, company, month, year)

        # ── Stream to response ────────────────────────────────────────────────
        month_name = calendar.month_abbr[month].upper() if month > 0 else "FULL_YEAR"
        filename   = f"GSTR1_{month_name}_{year}.xlsx"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"]      = buffer.getbuffer().nbytes
        return response

    # ── Sheet builders ────────────────────────────────────────────────────────

    def _build_gstr1_sheet(self, ws, rows, company, month, year):
        from openpyxl.styles.borders import Border, Side
        ws.title = "GST_DETAIL"

        # ── Title block ───────────────────────────────────────────────────────
        # Row 1 to 5: merge columns B to U
        ws.merge_cells("B1:U1")
        ws["B1"] = company.get("company_name", "").upper()
        ws["B1"].font = Font(bold=True, name="Calibri", size=11)
        ws["B1"].alignment = self._CENTER

        ws.merge_cells("B2:U2")
        ws["B2"] = company.get("address", "")
        ws["B2"].font = Font(name="Calibri", size=11)
        ws["B2"].alignment = self._CENTER

        ws.merge_cells("B3:U3")
        ws["B3"] = f"GSTIN : {company.get('gstin', 'N/A')}"
        ws["B3"].font = Font(name="Calibri", size=11)
        ws["B3"].alignment = self._CENTER

        ws.merge_cells("B4:U4")
        ws["B4"] = ""
        ws["B4"].font = Font(name="Calibri", size=11)
        ws["B4"].alignment = self._CENTER

        if month > 0:
            last_day = calendar.monthrange(year, month)[1]
            period_text = f"GSTR1 DETAILS FOR THE PERIOD 01/{month:02d}/{year} TO {last_day:02d}/{month:02d}/{year}"
        else:
            period_text = f"GSTR1 DETAILS FOR THE PERIOD 01/01/{year} TO 31/12/{year}"

        ws.merge_cells("B5:U5")
        ws["B5"] = period_text
        ws["B5"].font = Font(bold=True, name="Calibri", size=11)
        ws["B5"].alignment = self._CENTER

        # Column A is a spacer with small width
        ws.column_dimensions["A"].width = 2.71

        # ── Headers (Row 6 and Row 7) ─────────────────────────────────────────
        headers_r6 = {
            "B6": "S.No",
            "C6": "Desc",
            "D6": "GSTIN",
            "E6": "Invoice Date",
            "F6": "Invoice No.",
            "G6": "Invoice Value",
            "H6": "Local/\nCentral",
            "I6": "Invoice Type",
            "J6": "HSN Code",
            "K6": "Quantity",
            "L6": "Amount",
            "M6": "Taxable Amount",
            "N6": "SGST",
            "P6": "CGST",
            "R6": "IGST",
            "T6": "Cess",
            "U6": "Total GST"
        }
        for coord, text in headers_r6.items():
            ws[coord] = text

        headers_r7 = {
            "N7": "%age",
            "O7": "Amount",
            "P7": "%age",
            "Q7": "Amount",
            "R7": "%age",
            "S7": "Amount"
        }
        for coord, text in headers_r7.items():
            ws[coord] = text

        # Merge vertically or horizontally
        merges = [
            "B6:B7", "C6:C7", "D6:D7", "E6:E7", "F6:F7", "G6:G7",
            "H6:H7", "I6:I7", "J6:J7", "K6:K7", "L6:L7", "M6:M7",
            "N6:O6", "P6:Q6", "R6:S6", "T6:T7", "U6:U7"
        ]
        for merge_range in merges:
            ws.merge_cells(merge_range)

        # Style headers
        thin_side = Side(style='thin', color='000000')
        header_border = Border(top=thin_side, bottom=thin_side)
        header_font = Font(name="Calibri", size=11, bold=False)

        for r in [6, 7]:
            for c in range(2, 22): # Columns B to U
                cell = ws.cell(r, c)
                cell.font = header_font
                cell.border = header_border
                if c == 8: # Local/Central column wrap
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                else:
                    cell.alignment = self._CENTER

        # ── Group Rows by Section ─────────────────────────────────────────────
        b2b_list = []
        b2c_large_list = []
        b2c_small_list = []
        nil_rated_list = []

        for row in rows:
            sgst_pct = float(row.get("sgst_pct") or 0)
            cgst_pct = float(row.get("cgst_pct") or 0)
            igst_pct = float(row.get("igst_pct") or 0)
            rate = sgst_pct + cgst_pct + igst_pct

            is_b2b = bool(row.get("customer_gstin"))
            is_nil_rated = (rate == 0)

            if is_nil_rated:
                nil_rated_list.append(row)
            elif is_b2b:
                b2b_list.append(row)
            else:
                is_central = row.get("local_central") == "Central"
                grand_total = float(row.get("grand_total") or 0)
                if is_central and grand_total > 250000:
                    b2c_large_list.append(row)
                else:
                    b2c_small_list.append(row)

        sections = [
            (" B2B", b2b_list),
            (" B2C (Large) Invoice", b2c_large_list),
            (" B2C (Small) Invoice", b2c_small_list),
            (" Nil Rated/Exempted", nil_rated_list)
        ]

        current_row = 8
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11, bold=False)

        for title, items in sections:
            # Write Section Summary Row
            ws.cell(current_row, 1, "") # Col A
            ws.cell(current_row, 2, "") # Col B
            ws.cell(current_row, 3, title) # Col C

            if not items:
                # empty section spacing matching template (columns K to U as None)
                for c in range(4, 22):
                    ws.cell(current_row, c, "")
                for c in range(2, 22):
                    ws.cell(current_row, c).font = bold_font
                current_row += 1
                continue

            sum_qty = sum(float(item.get("quantity") or 0) for item in items)
            # Sum unique invoice values for Col L
            seen_invs = set()
            sum_invoice_value = 0.0
            for item in items:
                inv_no = item.get("invoice_number")
                if inv_no not in seen_invs:
                    seen_invs.add(inv_no)
                    sum_invoice_value += float(item.get("grand_total") or 0)

            sum_taxable = sum(float(item.get("taxable_value") or 0) for item in items)
            sum_sgst = sum(float(item.get("sgst_amount") or 0) for item in items)
            sum_cgst = sum(float(item.get("cgst_amount") or 0) for item in items)
            sum_igst = sum(float(item.get("igst_amount") or 0) for item in items)
            sum_cess = sum(float(item.get("cess") or 0) for item in items)
            sum_gst = sum(float(item.get("total_gst") or 0) for item in items)

            ws.cell(current_row, 11, sum_qty) # Col K (11)
            ws.cell(current_row, 12, sum_invoice_value) # Col L (12)
            ws.cell(current_row, 13, sum_taxable) # Col M (13)
            ws.cell(current_row, 14, 0.0) # Col N (14)
            ws.cell(current_row, 15, sum_sgst) # Col O (15)
            ws.cell(current_row, 16, 0.0) # Col P (16)
            ws.cell(current_row, 17, sum_cgst) # Col Q (17)
            ws.cell(current_row, 18, 0.0) # Col R (18)
            ws.cell(current_row, 19, sum_igst) # Col S (19)
            ws.cell(current_row, 20, sum_cess) # Col T (20)
            ws.cell(current_row, 21, sum_gst) # Col U (21)

            for c in range(2, 22):
                cell = ws.cell(current_row, c)
                cell.font = bold_font
                if c in [11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]:
                    cell.number_format = "#,##0.00"
                    cell.alignment = self._RIGHT

            current_row += 1

            # Write details
            for idx, item in enumerate(items, 1):
                sno_text = f"    {idx}."

                date_val = item.get("invoice_date")
                if date_val:
                    if hasattr(date_val, "strftime"):
                        date_str = date_val.strftime("%d/%m/%Y")
                    else:
                        try:
                            from datetime import datetime
                            date_str = datetime.strptime(str(date_val), "%Y-%m-%d").strftime("%d/%m/%Y")
                        except Exception:
                            date_str = str(date_val)
                else:
                    date_str = ""

                ws.cell(current_row, 1, "") # Col A
                ws.cell(current_row, 2, sno_text) # Col B
                ws.cell(current_row, 3, item.get("customer_name")) # Col C
                ws.cell(current_row, 4, item.get("customer_gstin") or "") # Col D
                ws.cell(current_row, 5, date_str) # Col E
                ws.cell(current_row, 6, item.get("invoice_number")) # Col F
                ws.cell(current_row, 7, float(item.get("grand_total") or 0)) # Col G
                ws.cell(current_row, 8, item.get("local_central")) # Col H
                ws.cell(current_row, 9, item.get("invoice_type")) # Col I
                ws.cell(current_row, 10, item.get("hsn_code") or "") # Col J
                ws.cell(current_row, 11, float(item.get("quantity") or 0)) # Col K
                ws.cell(current_row, 12, float(item.get("amount") or 0)) # Col L
                ws.cell(current_row, 13, float(item.get("taxable_value") or 0)) # Col M
                ws.cell(current_row, 14, float(item.get("sgst_pct") or 0)) # Col N
                ws.cell(current_row, 15, float(item.get("sgst_amount") or 0)) # Col O
                ws.cell(current_row, 16, float(item.get("cgst_pct") or 0)) # Col P
                ws.cell(current_row, 17, float(item.get("cgst_amount") or 0)) # Col Q
                ws.cell(current_row, 18, float(item.get("igst_pct") or 0)) # Col R
                ws.cell(current_row, 19, float(item.get("igst_amount") or 0)) # Col S
                ws.cell(current_row, 20, float(item.get("cess") or 0)) # Col T
                ws.cell(current_row, 21, float(item.get("total_gst") or 0)) # Col U

                for c in range(2, 22):
                    cell = ws.cell(current_row, c)
                    cell.font = regular_font
                    if c in [2, 4, 5, 6, 8, 9, 10]:
                        cell.alignment = self._CENTER
                    if c in [7, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]:
                        cell.number_format = "#,##0.00"
                        cell.alignment = self._RIGHT

                current_row += 1

        # ── Column widths ─────────────────────────────────────────────────────
        widths = {
            "A": 2.71, "B": 5.71, "C": 30.71, "D": 11.71, "E": 13.0, "F": 13.0, "G": 13.0,
            "H": 13.0, "I": 13.0, "J": 8.71, "K": 12.71, "L": 11.71, "M": 13.0, "N": 5.71,
            "O": 10.71, "P": 5.71, "Q": 10.71, "R": 5.71, "S": 10.71, "T": 13.0, "U": 13.0
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _build_hsn_sheet(self, ws, rows, company, month, year):
        month_label = f"{calendar.month_name[month]} {year}" if month > 0 else f"Full Year {year}"

        # ── Title block ───────────────────────────────────────────────────────
        ws.merge_cells("A1:G1")
        ws["A1"] = company.get("company_name", "").upper()
        ws["A1"].font      = Font(bold=True, size=13)
        ws["A1"].alignment = self._CENTER

        ws.merge_cells("A2:G2")
        ws["A2"] = f"GSTIN: {company.get('gstin', 'N/A')}"
        ws["A2"].font      = Font(size=10)
        ws["A2"].alignment = self._CENTER

        ws.merge_cells("A3:G3")
        ws["A3"] = f"HSN Summary — {month_label}"
        ws["A3"].font      = self._SECTION_FONT
        ws["A3"].fill      = self._SECTION_FILL
        ws["A3"].alignment = self._CENTER

        ws.append([])  # blank row 4

        # ── Headers ───────────────────────────────────────────────────────────
        headers = [
            "HSN Code", "Total Qty",
            "Taxable Value", "CGST", "SGST", "IGST", "Total",
        ]
        ws.append(headers)
        for cell in ws[5]:
            cell.font      = self._HEADER_FONT
            cell.fill      = self._HEADER_FILL
            cell.alignment = self._CENTER

        # ── Data rows ─────────────────────────────────────────────────────────
        for row in rows:
            ws.append([
                row.get("hsn_code") or "N/A",
                float(row.get("total_qty")     or 0),
                float(row.get("taxable_value") or 0),
                float(row.get("cgst")          or 0),
                float(row.get("sgst")          or 0),
                float(row.get("igst")          or 0),
                float(row.get("total")         or 0),
            ])
            for cell in ws[ws.max_row]:
                cell.alignment = self._CENTER
            for col in range(2, 8):
                ws.cell(ws.max_row, col).number_format = "#,##0.00"
                ws.cell(ws.max_row, col).alignment     = self._RIGHT

        # ── Totals row ────────────────────────────────────────────────────────
        if rows:
            total_row = ws.max_row + 1
            ws.cell(total_row, 1, "TOTAL")
            ws.cell(total_row, 1).font = self._TOTAL_FONT
            for col_idx, key in enumerate(
                ["total_qty", "taxable_value", "cgst", "sgst", "igst", "total"],
                start=2
            ):
                val = sum(float(r.get(key) or 0) for r in rows)
                cell = ws.cell(total_row, col_idx, val)
                cell.font          = self._TOTAL_FONT
                cell.fill          = self._TOTAL_FILL
                cell.number_format = "#,##0.00"
                cell.alignment     = self._RIGHT

        # ── Column widths ─────────────────────────────────────────────────────
        for col, width in zip("ABCDEFG", [14, 12, 16, 12, 12, 12, 14]):
            ws.column_dimensions[col].width = width